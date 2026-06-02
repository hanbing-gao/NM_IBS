from __future__ import annotations

import html
import argparse
import re
import zipfile
from pathlib import Path

import pandas as pd


PACKAGE_ROOT = Path(__file__).resolve().parents[1]
INPUT_FILE = PACKAGE_ROOT / "results" / "subtype_sensitivity" / "subtype_sensitivity_results_all_idps.csv"
OUT_DIR = PACKAGE_ROOT / "results" / "subtype_sensitivity"
DOCX_TEMPLATE: Path | None = None
ROI_PATH = PACKAGE_ROOT / "docs" / "ROI_IBS.csv"
MODALITY_ORDER = ["CT", "SA", "CV"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create formatted supplementary subtype-sensitivity tables from p6_subtype_sensitivity.py outputs."
    )
    parser.add_argument("--input-file", type=Path, default=INPUT_FILE)
    parser.add_argument("--output-dir", type=Path, default=OUT_DIR)
    parser.add_argument("--docs-dir", type=Path, default=PACKAGE_ROOT / "docs")
    parser.add_argument(
        "--docx-template",
        type=Path,
        default=None,
        help="Optional .docx template used for the short paste-ready summary document.",
    )
    return parser.parse_args()


def configure_paths(args: argparse.Namespace) -> None:
    global INPUT_FILE, OUT_DIR, DOCX_TEMPLATE, ROI_PATH
    INPUT_FILE = args.input_file.resolve()
    OUT_DIR = args.output_dir.resolve()
    DOCX_TEMPLATE = args.docx_template.resolve() if args.docx_template is not None else None
    ROI_PATH = args.docs_dir.resolve() / "ROI_IBS.csv"


def load_roi_order(df: pd.DataFrame) -> list[str]:
    if ROI_PATH.exists():
        return pd.read_csv(ROI_PATH)["ROI"].tolist()
    return df["roi"].drop_duplicates().tolist()


def sort_by_modality_roi(df: pd.DataFrame, roi_order: list[str]) -> pd.DataFrame:
    out = df.copy()
    out["_modality_order"] = pd.Categorical(
        out["modality"], categories=MODALITY_ORDER, ordered=True
    )
    out["_roi_order"] = pd.Categorical(out["roi"], categories=roi_order, ordered=True)
    out = out.sort_values(["_modality_order", "_roi_order"])
    out = out.drop(columns=["_modality_order", "_roi_order"])
    return out.reset_index(drop=True)


def _safe_sheet_name(name: str) -> str:
    return re.sub(r"[\[\]\:\*\?/\\]", "_", name)[:31]


def _format_excel(path: Path) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
    except Exception:
        return

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for column_cells in sheet.columns:
            max_len = 0
            letter = column_cells[0].column_letter
            for cell in column_cells[:200]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 45))
            sheet.column_dimensions[letter].width = max(10, max_len + 2)
    workbook.save(path)


def write_xlsx(path: Path, readme: list[tuple[str, str]], results: pd.DataFrame) -> None:
    readme_df = pd.DataFrame(readme, columns=["Item", "Description"])
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        readme_df.to_excel(writer, sheet_name="README", index=False)
        results.to_excel(writer, sheet_name=_safe_sheet_name("Results"), index=False)
    _format_excel(path)


def common_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df[["modality", "roi", "idp", "n_total", "n_hc", "n_ibs"]].copy()
    return out


def add_significance(table: pd.DataFrame, p_fdr_col: str) -> pd.DataFrame:
    table["significant_fdr05"] = table[p_fdr_col] < 0.05
    return table


def build_tables(df: pd.DataFrame) -> dict[str, tuple[pd.DataFrame, list[tuple[str, str]]]]:
    n_total = df["n_total"]
    df = df.copy()
    roi_order = load_roi_order(df)
    df["df_den_group5"] = n_total - 6
    df["df_den_group5_sex"] = n_total - 10
    df["df_den_nested"] = n_total - 6
    df["df_den_nested_sex"] = n_total - 10
    df["group5_omnibus_df_num"] = 4
    df["group5_sex_interaction_df_num"] = df["group5_sex_interaction_df"]
    df["nested_subtype_heterogeneity_df_num"] = df["nested_subtype_heterogeneity_df"]
    df["nested_subtype_sex_heterogeneity_df_num"] = df[
        "nested_subtype_heterogeneity_with_sex_interaction_df"
    ]

    group5_cols = [
        "modality",
        "roi",
        "idp",
        "n_total",
        "n_hc",
        "n_ibs",
        "group5_omnibus_df_num",
        "df_den_group5",
        "group5_omnibus_F",
        "group5_omnibus_p",
        "group5_omnibus_p_fdr",
        "adj_mean_HC",
        "adj_mean_IBS-C",
        "adj_mean_diff_IBS-C_minus_HC",
        "IBS-C_vs_HC_p_fdr",
        "adj_mean_IBS-D",
        "adj_mean_diff_IBS-D_minus_HC",
        "IBS-D_vs_HC_p_fdr",
        "adj_mean_IBS-M",
        "adj_mean_diff_IBS-M_minus_HC",
        "IBS-M_vs_HC_p_fdr",
        "adj_mean_IBS-U",
        "adj_mean_diff_IBS-U_minus_HC",
        "IBS-U_vs_HC_p_fdr",
    ]
    group5 = df[group5_cols].rename(
        columns={
            "group5_omnibus_df_num": "df_num",
            "df_den_group5": "df_den",
            "group5_omnibus_F": "F",
            "group5_omnibus_p": "p",
            "group5_omnibus_p_fdr": "p_fdr_within_modality",
        }
    )
    group5 = add_significance(group5, "p_fdr_within_modality")
    group5 = sort_by_modality_roi(group5, roi_order)

    group5_sex_cols = [
        "modality",
        "roi",
        "idp",
        "n_total",
        "n_hc",
        "n_ibs",
        "group5_sex_interaction_df_num",
        "df_den_group5_sex",
        "group5_sex_interaction_F",
        "group5_sex_interaction_p",
        "group5_sex_interaction_p_fdr",
    ]
    group5_sex = df[group5_sex_cols].rename(
        columns={
            "group5_sex_interaction_df_num": "df_num",
            "df_den_group5_sex": "df_den",
            "group5_sex_interaction_F": "F",
            "group5_sex_interaction_p": "p",
            "group5_sex_interaction_p_fdr": "p_fdr_within_modality",
        }
    )
    group5_sex = add_significance(group5_sex, "p_fdr_within_modality")
    group5_sex = sort_by_modality_roi(group5_sex, roi_order)

    nested_cols = [
        "modality",
        "roi",
        "idp",
        "n_total",
        "n_hc",
        "n_ibs",
        "main_ibs_beta",
        "main_ibs_p",
        "main_ibs_p_fdr",
        "nested_subtype_heterogeneity_df_num",
        "df_den_nested",
        "nested_subtype_heterogeneity_F",
        "nested_subtype_heterogeneity_p",
        "nested_subtype_heterogeneity_p_fdr",
        "adj_mean_HC",
        "adj_mean_diff_IBS-C_minus_HC",
        "adj_mean_diff_IBS-D_minus_HC",
        "adj_mean_diff_IBS-M_minus_HC",
        "adj_mean_diff_IBS-U_minus_HC",
    ]
    nested = df[nested_cols].rename(
        columns={
            "nested_subtype_heterogeneity_df_num": "df_num",
            "df_den_nested": "df_den",
            "nested_subtype_heterogeneity_F": "F",
            "nested_subtype_heterogeneity_p": "p",
            "nested_subtype_heterogeneity_p_fdr": "p_fdr_within_modality",
        }
    )
    nested = add_significance(nested, "p_fdr_within_modality")
    nested = sort_by_modality_roi(nested, roi_order)

    nested_sex_cols = [
        "modality",
        "roi",
        "idp",
        "n_total",
        "n_hc",
        "n_ibs",
        "common_group_sex_interaction_F",
        "common_group_sex_interaction_p",
        "common_group_sex_interaction_p_fdr",
        "nested_subtype_sex_heterogeneity_df_num",
        "df_den_nested_sex",
        "nested_subtype_heterogeneity_with_sex_interaction_F",
        "nested_subtype_heterogeneity_with_sex_interaction_p",
        "nested_subtype_heterogeneity_with_sex_interaction_p_fdr",
    ]
    nested_sex = df[nested_sex_cols].rename(
        columns={
            "nested_subtype_sex_heterogeneity_df_num": "df_num",
            "df_den_nested_sex": "df_den",
            "nested_subtype_heterogeneity_with_sex_interaction_F": "F",
            "nested_subtype_heterogeneity_with_sex_interaction_p": "p",
            "nested_subtype_heterogeneity_with_sex_interaction_p_fdr": "p_fdr_within_modality",
        }
    )
    nested_sex = add_significance(nested_sex, "p_fdr_within_modality")
    nested_sex = sort_by_modality_roi(nested_sex, roi_order)

    return {
        "01_five_level_group_omnibus": (
            group5,
            [
                ("Analysis", "Five-level group omnibus subtype sensitivity analysis."),
                (
                    "Model",
                    "Deviation score ~ five-level group + sex, with group coded as HC, IBS-C, IBS-D, IBS-M, IBS-U; HC is the reference level.",
                ),
                (
                    "Test",
                    "Omnibus F test of the five-level group term, assessing whether any IBS subtype differs from HC.",
                ),
                ("FDR", "Benjamini-Hochberg FDR correction was applied within each modality."),
            ],
        ),
        "02_five_level_group_by_sex_interaction": (
            group5_sex,
            [
                ("Analysis", "Five-level group-by-sex interaction subtype sensitivity analysis."),
                (
                    "Model",
                    "Deviation score ~ five-level group * sex, with group coded as HC, IBS-C, IBS-D, IBS-M, IBS-U; HC is the reference level.",
                ),
                (
                    "Test",
                    "Omnibus F test of the group-by-sex interaction, assessing sex-specific subtype effects.",
                ),
                ("FDR", "Benjamini-Hochberg FDR correction was applied within each modality."),
            ],
        ),
        "03_nested_subtype_heterogeneity": (
            nested,
            [
                ("Analysis", "Nested subtype heterogeneity comparison."),
                (
                    "Models compared",
                    "Binary common-effect model in which all IBS subtypes share one IBS effect versus a five-level group model allowing separate IBS-C, IBS-D, IBS-M, and IBS-U effects.",
                ),
                (
                    "Test",
                    "Nested F comparison testing whether allowing subtype-specific effects improves model fit beyond a common IBS effect.",
                ),
                ("FDR", "Benjamini-Hochberg FDR correction was applied within each modality."),
            ],
        ),
        "04_nested_subtype_by_sex_heterogeneity": (
            nested_sex,
            [
                ("Analysis", "Nested subtype-by-sex heterogeneity comparison."),
                (
                    "Models compared",
                    "Common IBS-by-sex interaction model versus a five-level group-by-sex model allowing separate subtype-by-sex interaction terms.",
                ),
                (
                    "Test",
                    "Nested F comparison testing whether sex-specific subtype effects improve model fit beyond a common IBS-by-sex interaction.",
                ),
                ("FDR", "Benjamini-Hochberg FDR correction was applied within each modality."),
            ],
        ),
    }


def build_summary(tables: dict[str, tuple[pd.DataFrame, list[tuple[str, str]]]]) -> pd.DataFrame:
    labels = {
        "01_five_level_group_omnibus": "Five-level group omnibus",
        "02_five_level_group_by_sex_interaction": "Five-level group-by-sex interaction",
        "03_nested_subtype_heterogeneity": "Nested subtype heterogeneity",
        "04_nested_subtype_by_sex_heterogeneity": "Nested subtype-by-sex heterogeneity",
    }
    rows = []
    for key, (table, _) in tables.items():
        for modality, sub in table.groupby("modality", sort=True):
            top = sub.sort_values("p").iloc[0]
            rows.append(
                {
                    "analysis": labels[key],
                    "modality": modality,
                    "n_idp": int(len(sub)),
                    "n_fdr_significant": int(sub["significant_fdr05"].sum()),
                    "min_raw_p": float(sub["p"].min()),
                    "min_fdr_p": float(sub["p_fdr_within_modality"].min()),
                    "top_roi_by_raw_p": top["roi"],
                    "top_roi_raw_p": float(top["p"]),
                    "top_roi_fdr_p": float(top["p_fdr_within_modality"]),
                }
            )
    return pd.DataFrame(rows)


def esc(text: object) -> str:
    return html.escape(str(text), quote=False)


def p(text: str, style: str | None = None, bold: bool = False) -> str:
    ppr = ""
    if style:
        ppr = f'<w:pPr><w:pStyle w:val="{style}"/><w:spacing w:after="120"/></w:pPr>'
    else:
        ppr = '<w:pPr><w:spacing w:after="120"/></w:pPr>'
    rpr = "<w:rPr><w:b/></w:rPr>" if bold else ""
    return f'<w:p>{ppr}<w:r>{rpr}<w:t xml:space="preserve">{esc(text)}</w:t></w:r></w:p>'


def table_xml(df: pd.DataFrame) -> str:
    rows = []
    rows.append(list(df.columns))
    for _, row in df.iterrows():
        rows.append([row[col] for col in df.columns])
    xml = [
        "<w:tbl><w:tblPr><w:tblW w:w=\"0\" w:type=\"auto\"/>"
        "<w:tblBorders>"
        '<w:top w:val="single" w:sz="4" w:space="0" w:color="BFBFBF"/>'
        '<w:left w:val="single" w:sz="4" w:space="0" w:color="BFBFBF"/>'
        '<w:bottom w:val="single" w:sz="4" w:space="0" w:color="BFBFBF"/>'
        '<w:right w:val="single" w:sz="4" w:space="0" w:color="BFBFBF"/>'
        '<w:insideH w:val="single" w:sz="4" w:space="0" w:color="BFBFBF"/>'
        '<w:insideV w:val="single" w:sz="4" w:space="0" w:color="BFBFBF"/>'
        "</w:tblBorders></w:tblPr>"
    ]
    for row_i, row in enumerate(rows):
        xml.append("<w:tr>")
        for value in row:
            display = value
            if isinstance(value, float):
                display = f"{value:.4g}"
            props = '<w:shd w:val="clear" w:color="auto" w:fill="EAF3F8"/>' if row_i == 0 else ""
            bold = "<w:rPr><w:b/></w:rPr>" if row_i == 0 else ""
            xml.append(
                f"<w:tc><w:tcPr>{props}</w:tcPr><w:p><w:r>{bold}<w:t>{esc(display)}</w:t></w:r></w:p></w:tc>"
            )
        xml.append("</w:tr>")
    xml.append("</w:tbl>")
    return "".join(xml)


def write_summary_docx(path: Path, summary: pd.DataFrame) -> None:
    if DOCX_TEMPLATE is None or not DOCX_TEMPLATE.exists():
        path.with_suffix(".md").write_text(
            "# IBS subtype sensitivity analyses\n\n"
            "A .docx template was not provided, so this run wrote the formatted Excel/CSV "
            "tables only. The summary table is available in the companion CSV/XLSX files.\n",
            encoding="utf-8",
        )
        return

    def min_fdr_fragment(label: str) -> str:
        sub = summary[summary["analysis"] == label].set_index("modality")
        parts = []
        for modality in ["CT", "SA", "CV"]:
            parts.append(f"{modality} = {sub.loc[modality, 'min_fdr_p']:.3f}")
        return ", ".join(parts)

    result_sentence = (
        "Across cortical thickness (CT), surface area (SA), and cortical volume (CV), no "
        "five-level group omnibus effect, five-level group-by-sex interaction, nested subtype "
        "heterogeneity effect, or nested subtype-by-sex heterogeneity effect survived FDR "
        "correction within modality."
    )
    methods_text = (
        "IBS subtype sensitivity analyses were conducted to evaluate whether heterogeneity "
        "across IBS subtypes could mask an overall IBS-HC effect. IBS cases were classified "
        "as IBS-C, IBS-D, IBS-M, or IBS-U according to Rome III criteria using DHQ-derived "
        "bowel-habit information. For each cortical measure, models included sex and coded "
        "group as HC, IBS-C, IBS-D, IBS-M, and IBS-U, with HC as the reference group. We tested "
        "the omnibus five-level group effect, the group-by-sex interaction, nested subtype "
        "heterogeneity beyond a common IBS effect, and nested subtype-by-sex heterogeneity beyond "
        "a common IBS-by-sex interaction. FDR correction was performed separately within each "
        "modality for each analysis."
    )
    results_text = (
        "No analysis yielded FDR-significant subtype-related effects. The minimum FDR-corrected "
        "p values across modalities were: five-level group omnibus, "
        f"{min_fdr_fragment('Five-level group omnibus')}; five-level group-by-sex interaction, "
        f"{min_fdr_fragment('Five-level group-by-sex interaction')}; nested subtype heterogeneity, "
        f"{min_fdr_fragment('Nested subtype heterogeneity')}; and nested subtype-by-sex heterogeneity, "
        f"{min_fdr_fragment('Nested subtype-by-sex heterogeneity')}. These findings do not support the "
        "interpretation that DHQ-derived IBS subtype heterogeneity or sex-specific subtype effects "
        "masked robust average IBS-HC cortical morphometry differences."
    )
    caption_text = (
        "Supplementary Table. IBS subtype sensitivity analyses. The four accompanying Excel files "
        "report ROI-level statistics for the five-level subtype group omnibus effect, the five-level "
        "group-by-sex interaction, the nested subtype heterogeneity comparison, and the nested "
        "subtype-by-sex heterogeneity comparison. P values were FDR-corrected separately within "
        "CT, SA, and CV for each analysis."
    )
    summary_table = summary[
        [
            "analysis",
            "modality",
            "n_idp",
            "n_fdr_significant",
            "min_fdr_p",
            "top_roi_by_raw_p",
            "top_roi_raw_p",
            "top_roi_fdr_p",
        ]
    ].copy()

    body = [
        p("IBS Subtype Sensitivity Analyses: Supplementary Table Summary", style="Title"),
        p("Paste-ready Methods text", style="Heading1"),
        p(methods_text),
        p("Paste-ready Results text", style="Heading1"),
        p(results_text),
        p("One-sentence summary", style="Heading1"),
        p(result_sentence),
        p("Suggested supplementary caption", style="Heading1"),
        p(caption_text),
        p("Summary by modality", style="Heading1"),
        table_xml(summary_table),
    ]
    sect = (
        '<w:sectPr><w:pgSz w:w="11906" w:h="16838"/>'
        '<w:pgMar w:top="1440" w:right="1440" w:bottom="1440" '
        'w:left="1440" w:header="708" w:footer="708" w:gutter="0"/>'
        '<w:cols w:space="708"/><w:docGrid w:linePitch="360"/></w:sectPr>'
    )
    document = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main" '
        'xmlns:xml="http://www.w3.org/XML/1998/namespace"><w:body>'
        + "".join(body)
        + sect
        + "</w:body></w:document>"
    )
    tmp = path.with_name(path.name + ".buildtmp")
    with zipfile.ZipFile(DOCX_TEMPLATE, "r") as zin:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == "word/document.xml":
                    zout.writestr(item, document.encode("utf-8"))
                else:
                    zout.writestr(item, zin.read(item.filename))
    with zipfile.ZipFile(tmp) as zf:
        bad = zf.testzip()
    if bad:
        raise RuntimeError(f"Bad docx member: {bad}")
    path.write_bytes(tmp.read_bytes())
    try:
        tmp.unlink()
    except OSError:
        pass


def main() -> None:
    args = parse_args()
    configure_paths(args)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.read_csv(INPUT_FILE)
    tables = build_tables(df)

    for key, (table, readme) in tables.items():
        out_xlsx = OUT_DIR / f"Supplementary_Table_SubtypeSensitivity_{key}.xlsx"
        out_csv = OUT_DIR / f"Supplementary_Table_SubtypeSensitivity_{key}.csv"
        write_xlsx(out_xlsx, readme, table)
        table.to_csv(out_csv, index=False)

    summary = build_summary(tables)
    summary_csv = OUT_DIR / "Supplementary_Table_SubtypeSensitivity_summary_by_modality.csv"
    summary_xlsx = OUT_DIR / "Supplementary_Table_SubtypeSensitivity_summary_by_modality.xlsx"
    summary.to_csv(summary_csv, index=False)
    write_xlsx(
        summary_xlsx,
        [
            (
                "Description",
                "Summary of subtype sensitivity analyses by effect and modality.",
            ),
            (
                "FDR",
                "Benjamini-Hochberg correction was applied within each modality and within each analysis.",
            ),
        ],
        summary,
    )

    write_summary_docx(
        OUT_DIR / "Supplementary_Table_SubtypeSensitivity_summary_and_captions.docx",
        summary,
    )

    print(f"Wrote supplementary subtype sensitivity tables to: {OUT_DIR}")
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
